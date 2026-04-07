"""
Generate Tabel Analisis Hasil Pengujian OMR
Untuk keperluan Skripsi
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Hasil Pengujian OMR"

# ============================================================
# STYLES
# ============================================================
def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    thin = Side(style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def make_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, name="Calibri", size=size)

def make_align(h="center", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

# Color palette
COLOR_HEADER_DARK  = "1F4E79"   # Biru tua (header utama)
COLOR_HEADER_MID   = "2E75B6"   # Biru sedang (sub-header)
COLOR_CAT_1        = "DEEAF1"   # Biru muda (kategori 1)
COLOR_CAT_2        = "E2EFDA"   # Hijau muda (kategori 2)
COLOR_CAT_3        = "FFF2CC"   # Kuning muda (kategori 3)
COLOR_ROW_ALT      = "F2F2F2"   # Abu strip
COLOR_STATUS_PASS  = "C6EFCE"   # Hijau hasil
COLOR_STATUS_WARN  = "FFEB9C"   # Kuning hasil
COLOR_STATUS_FAIL  = "FFC7CE"   # Merah hasil
COLOR_WHITE        = "FFFFFF"

# ============================================================
# JUDUL
# ============================================================
ws.merge_cells("A1:L1")
title_cell = ws["A1"]
title_cell.value = "Tabel Hasil Pengujian Sistem OMR (Optical Mark Recognition)"
title_cell.font = Font(bold=True, size=13, name="Calibri", color=COLOR_WHITE)
title_cell.fill = make_fill(COLOR_HEADER_DARK)
title_cell.alignment = make_align()
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:L2")
sub_cell = ws["A2"]
sub_cell.value = "Pengujian dilakukan menggunakan Document Scanner — Backend: FastAPI + OpenCV"
sub_cell.font = Font(italic=True, size=10, name="Calibri", color="595959")
sub_cell.fill = make_fill("D6E4F0")
sub_cell.alignment = make_align()
ws.row_dimensions[2].height = 18

# ============================================================
# HEADER KOLOM
# ============================================================
headers = [
    "No",
    "Kode\nTest Case",
    "Kategori",
    "Nama Skenario",
    "Kondisi LJK",
    "Kelemahan\nScanner Terkait",
    "Ekspektasi\nSistem",
    "Jumlah Soal\nBenar",
    "Jumlah Soal\nSalah",
    "Jumlah\nKosong",
    "Akurasi (%)",
    "Status\nHasil",
]

header_row = 4
ws.row_dimensions[header_row] = ws.row_dimensions[header_row]
ws.row_dimensions[header_row].height = 45

for col_idx, header_text in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col_idx)
    cell.value = header_text
    cell.font = Font(bold=True, size=10, name="Calibri", color=COLOR_WHITE)
    cell.fill = make_fill(COLOR_HEADER_MID)
    cell.alignment = make_align()
    cell.border = make_border()

# Kosongkan baris 3 sbg spacer
ws.row_dimensions[3].height = 8
ws.merge_cells("A3:L3")
ws["A3"].fill = make_fill("FFFFFF")

# ============================================================
# DATA TEST CASES
# ============================================================
#   No, Kode, Kategori, Nama, Kondisi, Kelemahan Scanner, Ekspektasi,
#   Benar, Salah, Kosong, Akurasi, Status
# Field Benar/Salah/Kosong/Akurasi/Status dikosongkan (diisi saat pengujian)

data = [
    # --- KATEGORI 1 ---
    (1,  "TC-01", "Kondisi Fisik Kertas",
     "Ideal Condition (Baseline)",
     "Kertas mulus, pensil 2B, pencahayaan merata, isian bulat penuh",
     "Tidak ada (kondisi sempurna)",
     "Semua 30 soal terdeteksi benar; Akurasi = 100%",
     "", "", "", "", ""),

    (2,  "TC-02", "Kondisi Fisik Kertas",
     "Kertas Miring (Skew)",
     "LJK dimasukkan ke scanner posisi miring 5–15 derajat",
     "ADF scanner sering menarik kertas tidak lurus (mechanical skew)",
     "Perspective warp berhasil meluruskan grid; Akurasi minimal 90%",
     "", "", "", "", ""),

    (3,  "TC-03", "Kondisi Fisik Kertas",
     "Kertas Terlipat / Lecek (Crease)",
     "LJK dilipat menjadi 4, diratakan kembali lalu di-scan",
     "Lipatan menciptakan bayangan garis hitam di hasil scan",
     "Line removal menghapus bayangan; Akurasi minimal 85%",
     "", "", "", "", ""),

    # --- KATEGORI 2 ---
    (4,  "TC-04", "Masalah Optikal Scanner",
     "Bleed-Through (Tembus Pandang)",
     "Tulisan tebal di bagian belakang LJK, scan dengan brightness normal",
     "Lampu exposure terlalu kuat membuat tinta halaman belakang tembus",
     "Thresholding memisahkan noise bleed-through; Akurasi minimal 80%",
     "", "", "", "", ""),

    (5,  "TC-05", "Masalah Optikal Scanner",
     "Over-Exposed / Isian Pensil Tipis",
     "Isian pensil tipis/tekanan ringan, scanner setting kontras tinggi",
     "Mode 'Document Text' scanner sering wash-out arsiran pensil tipis",
     "Sistem masih mendeteksi isian tipis; cek MIN_FILL_THRESHOLD",
     "", "", "", "", ""),

    (6,  "TC-06", "Masalah Optikal Scanner",
     "Sensor Dust / Garis Debu Scanner",
     "Debu di kaca scanner; simulasi: garis vertikal tipis menembus grid",
     "Debu di kaca ADF menghasilkan garis vertikal hitam sepanjang kertas",
     "Vertical line removal membersihkan garis debu tanpa mempengaruhi bubble",
     "", "", "", "", ""),

    (7,  "TC-07", "Masalah Optikal Scanner",
     "Marker Sudut Terpotong (Auto-Crop Error)",
     "Fitur Auto-Crop scanner memotong ≥50% salah satu marker sudut",
     "Auto-Crop bawaan scanner sering terlalu agresif memotong tepi putih",
     "Sistem mengeluarkan error graceful: 'Kertas LJK tidak terdeteksi'",
     "", "", "", "", ""),

    # --- KATEGORI 3 ---
    (8,  "TC-08", "Perilaku Pengisian Siswa",
     "Hapusan Kotor (Ghost Marks / Smudge)",
     "Jawaban salah dihapus tidak bersih; ada sisa bayangan abu-abu di opsi A, isian penuh di opsi B",
     "Scanner menangkap sisa hapusan sebagai coretan abu-abu solid",
     "Sistem memilih jawaban dengan piksel terbanyak; sisa hapusan diabaikan",
     "", "", "", "", ""),

    (9,  "TC-09", "Perilaku Pengisian Siswa",
     "Gaya Isian Tidak Standar (Tanda X / Centang)",
     "Siswa mengisi dengan tanda silang (X) atau centang (✓) bukan bulatan penuh",
     "-",
     "Sistem masih mendeteksi tanda jika melebihi MIN_FILL_THRESHOLD",
     "", "", "", "", ""),

    (10, "TC-10", "Perilaku Pengisian Siswa",
     "Variasi Resolusi Scan (DPI Berbeda)",
     "Gambar LJK yang sama di-scan dengan 150 DPI, 300 DPI, dan 600 DPI",
     "Resolusi DPI yang berbeda menghasilkan ukuran file sangat berbeda",
     "Setelah warp ke (1000×1414), akurasi konsisten di semua resolusi",
     "", "", "", "", ""),
]

# Warna per kategori
cat_colors = {
    "Kondisi Fisik Kertas":    COLOR_CAT_1,
    "Masalah Optikal Scanner": COLOR_CAT_2,
    "Perilaku Pengisian Siswa": COLOR_CAT_3,
}

for row_offset, row_data in enumerate(data):
    excel_row = header_row + 1 + row_offset
    ws.row_dimensions[excel_row].height = 55

    cat_name = row_data[2]
    bg_color = cat_colors.get(cat_name, COLOR_WHITE)
    # Alternating shade
    if row_offset % 2 == 1:
        # Sedikit lebih gelap, pakai strip
        pass  # tetap pakai cat_color

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=excel_row, column=col_idx)
        cell.value = value
        cell.border = make_border()
        cell.font = make_font(size=10)
        cell.alignment = make_align(h="left" if col_idx >= 4 else "center")

        # Background per kategori
        cell.fill = make_fill(bg_color)

        # Kolom No & Kode: bold center
        if col_idx in (1, 2):
            cell.font = make_font(bold=True, size=10)
            cell.alignment = make_align()

        # Kolom Status (kosong, siap diisi): highlight
        if col_idx == 12:
            cell.fill = make_fill(COLOR_ROW_ALT)
            cell.alignment = make_align()

# ============================================================
# LEGEND STATUS (di bawah tabel)
# ============================================================
legend_row = header_row + len(data) + 2
ws.merge_cells(f"A{legend_row}:C{legend_row}")
ws[f"A{legend_row}"].value = "KETERANGAN STATUS:"
ws[f"A{legend_row}"].font = make_font(bold=True, size=10)

items = [
    ("PASS", COLOR_STATUS_PASS, "Sistem berhasil memproses dan akurasi sesuai ekspektasi"),
    ("WARNING", COLOR_STATUS_WARN, "Sistem berhasil tapi akurasi di bawah ekspektasi / ada anomali minor"),
    ("FAIL", COLOR_STATUS_FAIL, "Sistem gagal memproses atau akurasi jauh di bawah ekspektasi"),
]
for i, (label, color, desc) in enumerate(items):
    r = legend_row + 1 + i
    ws[f"A{r}"].value = label
    ws[f"A{r}"].fill = make_fill(color)
    ws[f"A{r}"].font = make_font(bold=True)
    ws[f"A{r}"].alignment = make_align()
    ws[f"A{r}"].border = make_border()
    ws.merge_cells(f"B{r}:L{r}")
    ws[f"B{r}"].value = desc
    ws[f"B{r}"].font = make_font(size=10)
    ws[f"B{r}"].alignment = make_align(h="left")
    ws[f"B{r}"].border = make_border()

# ============================================================
# COLUMN WIDTHS
# ============================================================
col_widths = [5, 10, 22, 30, 45, 42, 45, 12, 12, 10, 12, 14]
for i, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ============================================================
# FREEZE PANES & SHEET SETUP
# ============================================================
ws.freeze_panes = f"A{header_row + 1}"  # Freeze header
ws.sheet_view.showGridLines = True
ws.print_title_rows = f"$1:${header_row}"  # Print header di tiap halaman

# Page setup untuk print
from openpyxl.worksheet.page import PageMargins
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

# ============================================================
# SAVE
# ============================================================
output_path = "Tabel_Pengujian_OMR.xlsx"
wb.save(output_path)
print(f"\n✅ File berhasil dibuat: {output_path}")
print(f"   Buka file tersebut, lalu isi kolom:")
print(f"   'Jumlah Soal Benar', 'Salah', 'Kosong', 'Akurasi', 'Status'")
print(f"   setelah kamu selesai pengujian.\n")
